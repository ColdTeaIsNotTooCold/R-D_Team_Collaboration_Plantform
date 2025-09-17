"""
文档处理引擎
负责文件解析、文本提取、分块处理、元数据提取等功能
"""

import os
import logging
import hashlib
from typing import Dict, List, Optional, Any, Tuple, BinaryIO
from pathlib import Path
from datetime import datetime
import json

import chromadb
from sentence_transformers import SentenceTransformer
import numpy as np

from .config import settings
from ..utils.text_utils import TextUtils

logger = logging.getLogger(__name__)


class DocumentProcessor:
    """文档处理引擎"""

    def __init__(self):
        """初始化文档处理器"""
        self.text_utils = TextUtils()
        self.chroma_client = None
        self.embedding_model = None
        self.collection = None

        # 支持的文件类型
        self.supported_extensions = {
            '.txt', '.md', '.py', '.js', '.html', '.css',
            '.json', '.xml', '.csv', '.log',
            '.pdf', '.docx', '.xlsx'
        }

        # 初始化向量数据库
        self._init_vector_db()

        # 初始化嵌入模型
        self._init_embedding_model()

    def _init_vector_db(self):
        """初始化ChromaDB向量数据库"""
        try:
            self.chroma_client = chromadb.PersistentClient(
                path=os.path.join(os.getcwd(), "chroma_db")
            )

            # 获取或创建集合
            self.collection = self.chroma_client.get_or_create_collection(
                name="documents",
                metadata={"description": "文档向量存储"}
            )

            logger.info("ChromaDB向量数据库初始化成功")
        except Exception as e:
            logger.error(f"初始化向量数据库失败: {e}")
            raise

    def _init_embedding_model(self):
        """初始化文本嵌入模型"""
        try:
            # 使用轻量级的中文嵌入模型
            self.embedding_model = SentenceTransformer(
                'paraphrase-multilingual-MiniLM-L12-v2'
            )
            logger.info("文本嵌入模型初始化成功")
        except Exception as e:
            logger.error(f"初始化嵌入模型失败: {e}")
            raise

    def process_document(self, file_path: str, metadata: Optional[Dict] = None) -> Dict[str, Any]:
        """
        处理单个文档

        Args:
            file_path: 文件路径
            metadata: 额外的元数据

        Returns:
            处理结果
        """
        try:
            # 验证文件
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"文件不存在: {file_path}")

            file_ext = Path(file_path).suffix.lower()
            if file_ext not in self.supported_extensions:
                raise ValueError(f"不支持的文件类型: {file_ext}")

            # 提取基础元数据
            base_metadata = self._extract_base_metadata(file_path)

            # 合并额外元数据
            if metadata:
                base_metadata.update(metadata)

            # 提取文本内容
            text_content = self._extract_text_content(file_path, file_ext)

            # 文本预处理
            cleaned_text = self.text_utils.clean_text(text_content)

            # 文本分块
            chunks = self.text_utils.chunk_text(cleaned_text)

            # 生成向量嵌入
            embeddings = self._generate_embeddings(chunks)

            # 存储到向量数据库
            doc_id = self._store_to_vector_db(
                file_path, chunks, embeddings, base_metadata
            )

            result = {
                "document_id": doc_id,
                "file_path": file_path,
                "metadata": base_metadata,
                "chunks_count": len(chunks),
                "total_length": len(cleaned_text),
                "processed_at": datetime.now().isoformat(),
                "status": "success"
            }

            logger.info(f"文档处理完成: {file_path}")
            return result

        except Exception as e:
            logger.error(f"处理文档失败 {file_path}: {e}")
            return {
                "file_path": file_path,
                "error": str(e),
                "status": "failed",
                "processed_at": datetime.now().isoformat()
            }

    def _extract_base_metadata(self, file_path: str) -> Dict[str, Any]:
        """提取基础元数据"""
        path = Path(file_path)

        # 计算文件哈希
        file_hash = self._calculate_file_hash(file_path)

        # 获取文件统计信息
        stat = path.stat()

        return {
            "file_name": path.name,
            "file_extension": path.suffix.lower(),
            "file_size": stat.st_size,
            "created_at": datetime.fromtimestamp(stat.st_ctime).isoformat(),
            "modified_at": datetime.fromtimestamp(stat.st_mtime).isoformat(),
            "file_hash": file_hash,
            "file_path": str(path.absolute())
        }

    def _calculate_file_hash(self, file_path: str) -> str:
        """计算文件哈希值"""
        hash_md5 = hashlib.md5()
        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_md5.update(chunk)
        return hash_md5.hexdigest()

    def _extract_text_content(self, file_path: str, file_ext: str) -> str:
        """根据文件类型提取文本内容"""
        try:
            if file_ext in ['.txt', '.md', '.py', '.js', '.html', '.css', '.json', '.xml', '.csv', '.log']:
                return self._extract_from_text_file(file_path)
            elif file_ext == '.pdf':
                return self._extract_from_pdf(file_path)
            elif file_ext == '.docx':
                return self._extract_from_docx(file_path)
            elif file_ext == '.xlsx':
                return self._extract_from_excel(file_path)
            else:
                raise ValueError(f"不支持的文件类型: {file_ext}")
        except Exception as e:
            logger.error(f"提取文本内容失败 {file_path}: {e}")
            raise

    def _extract_from_text_file(self, file_path: str) -> str:
        """从文本文件提取内容"""
        try:
            # 尝试UTF-8编码
            with open(file_path, 'r', encoding='utf-8') as f:
                return f.read()
        except UnicodeDecodeError:
            # 尝试GBK编码
            with open(file_path, 'r', encoding='gbk') as f:
                return f.read()

    def _extract_from_pdf(self, file_path: str) -> str:
        """从PDF文件提取文本"""
        try:
            import PyPDF2

            text_content = []
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page_num, page in enumerate(pdf_reader.pages):
                    try:
                        page_text = page.extract_text()
                        if page_text.strip():
                            text_content.append(f"=== 第{page_num + 1}页 ===\n{page_text}")
                    except Exception as e:
                        logger.warning(f"提取PDF第{page_num + 1}页失败: {e}")
                        continue

            return "\n\n".join(text_content)
        except ImportError:
            raise ImportError("PyPDF2未安装，无法处理PDF文件")
        except Exception as e:
            logger.error(f"PDF提取失败: {e}")
            raise

    def _extract_from_docx(self, file_path: str) -> str:
        """从Word文档提取文本"""
        try:
            from docx import Document

            doc = Document(file_path)
            text_content = []

            for paragraph in doc.paragraphs:
                if paragraph.text.strip():
                    text_content.append(paragraph.text)

            return "\n".join(text_content)
        except ImportError:
            raise ImportError("python-docx未安装，无法处理Word文档")
        except Exception as e:
            logger.error(f"Word文档提取失败: {e}")
            raise

    def _extract_from_excel(self, file_path: str) -> str:
        """从Excel文件提取文本"""
        try:
            import openpyxl

            workbook = openpyxl.load_workbook(file_path, read_only=True)
            text_content = []

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_text = [f"=== 工作表: {sheet_name} ==="]

                for row in sheet.iter_rows(values_only=True):
                    row_text = [str(cell) if cell is not None else "" for cell in row]
                    if any(cell.strip() for cell in row_text):
                        sheet_text.append("\t".join(row_text))

                if len(sheet_text) > 1:  # 只有有内容的工作表才添加
                    text_content.append("\n".join(sheet_text))

            workbook.close()
            return "\n\n".join(text_content)
        except ImportError:
            raise ImportError("openpyxl未安装，无法处理Excel文件")
        except Exception as e:
            logger.error(f"Excel文件提取失败: {e}")
            raise

    def _generate_embeddings(self, chunks: List[str]) -> List[List[float]]:
        """生成文本块的向量嵌入"""
        try:
            embeddings = []
            for chunk in chunks:
                if chunk.strip():
                    embedding = self.embedding_model.encode(chunk)
                    embeddings.append(embedding.tolist())
                else:
                    embeddings.append([0.0] * self.embedding_model.get_sentence_embedding_dimension())

            return embeddings
        except Exception as e:
            logger.error(f"生成向量嵌入失败: {e}")
            raise

    def _store_to_vector_db(self, file_path: str, chunks: List[str],
                           embeddings: List[List[float]], metadata: Dict[str, Any]) -> str:
        """存储到向量数据库"""
        try:
            # 生成文档ID
            doc_id = f"doc_{hashlib.md5(file_path.encode()).hexdigest()}"

            # 准备存储数据
            documents = []
            metadatas = []
            ids = []

            for i, (chunk, embedding) in enumerate(zip(chunks, embeddings)):
                chunk_metadata = metadata.copy()
                chunk_metadata.update({
                    "chunk_id": f"{doc_id}_chunk_{i}",
                    "chunk_index": i,
                    "chunk_length": len(chunk)
                })

                documents.append(chunk)
                metadatas.append(chunk_metadata)
                ids.append(f"{doc_id}_chunk_{i}")

            # 批量添加到向量数据库
            self.collection.add(
                documents=documents,
                embeddings=embeddings,
                metadatas=metadatas,
                ids=ids
            )

            return doc_id

        except Exception as e:
            logger.error(f"存储到向量数据库失败: {e}")
            raise

    def search_documents(self, query: str, limit: int = 10,
                        metadata_filter: Optional[Dict] = None) -> List[Dict[str, Any]]:
        """
        搜索文档

        Args:
            query: 搜索查询
            limit: 返回结果数量
            metadata_filter: 元数据过滤条件

        Returns:
            搜索结果
        """
        try:
            # 生成查询向量
            query_embedding = self.embedding_model.encode(query).tolist()

            # 构建查询条件
            query_conditions = {
                "query_embeddings": [query_embedding],
                "n_results": limit
            }

            if metadata_filter:
                query_conditions["where"] = metadata_filter

            # 执行搜索
            results = self.collection.query(**query_conditions)

            # 格式化结果
            formatted_results = []
            if results and results.get('documents'):
                for i, (doc, metadata, distance) in enumerate(zip(
                    results['documents'][0],
                    results['metadatas'][0],
                    results['distances'][0]
                )):
                    formatted_results.append({
                        "content": doc,
                        "metadata": metadata,
                        "distance": distance,
                        "similarity": 1 - distance  # 转换为相似度
                    })

            return formatted_results

        except Exception as e:
            logger.error(f"搜索文档失败: {e}")
            return []

    def get_document_by_id(self, doc_id: str) -> Optional[Dict[str, Any]]:
        """根据文档ID获取文档信息"""
        try:
            # 搜索特定文档的所有块
            results = self.collection.query(
                query_texts=[""],
                where={"document_id": doc_id},
                n_results=1000
            )

            if not results or not results.get('documents'):
                return None

            # 合并所有块的信息
            document_info = {
                "document_id": doc_id,
                "chunks": [],
                "total_chunks": len(results['documents'][0]),
                "metadata": results['metadatas'][0][0] if results['metadatas'] else {}
            }

            for doc, metadata in zip(results['documents'][0], results['metadatas'][0]):
                document_info["chunks"].append({
                    "content": doc,
                    "chunk_metadata": metadata
                })

            return document_info

        except Exception as e:
            logger.error(f"获取文档信息失败: {e}")
            return None

    def delete_document(self, doc_id: str) -> bool:
        """删除文档"""
        try:
            # 获取文档的所有块ID
            results = self.collection.query(
                query_texts=[""],
                where={"document_id": doc_id},
                n_results=1000
            )

            if results and results.get('ids'):
                # 删除所有块
                self.collection.delete(ids=results['ids'][0])
                logger.info(f"文档删除成功: {doc_id}")
                return True

            return False

        except Exception as e:
            logger.error(f"删除文档失败: {e}")
            return False

    def get_statistics(self) -> Dict[str, Any]:
        """获取文档处理统计信息"""
        try:
            # 获取集合信息
            collection_info = self.collection.count()

            return {
                "total_documents": collection_info,
                "supported_extensions": list(self.supported_extensions),
                "embedding_model": self.embedding_model.get_sentence_embedding_dimension(),
                "vector_db_path": os.path.join(os.getcwd(), "chroma_db")
            }

        except Exception as e:
            logger.error(f"获取统计信息失败: {e}")
            return {"error": str(e)}